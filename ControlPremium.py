"""
Filopinturas v4.0
Versi\u00f3n con Notebook (solapas): Clientes, Presupuestos, Stock, Proveedores, Movimientos (Ingreso/Egreso/Saldo)
Tema claro. Logo "Stoplc" en la parte superior de todas las solapas.

Requisitos:
- Python 3.x
- pip install pillow reportlab

Archivo: Filopinturas_v4.py
"""

import tkinter as tk
print("Iniciando Filopinturas v4.0 con nuevos botones...")
import openpyxl
from tkinter import ttk, messagebox, simpledialog, filedialog
from tkinter.filedialog import asksaveasfilename
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from PIL import Image, ImageTk
import sqlite3
import os
import shutil
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import webbrowser
import urllib.parse


# ---------------------
# CONFIG
# ---------------------
APP_TITLE = "Control Premium - v4.0"
# LOGO_PATH se define después de la inicialización de la DB
PDF_DIR = "presupuestos"
REPORT_DIR = "reportes"
DB_PATH = "controlpremium.db"

os.makedirs(PDF_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)

# ---------------------
# BASE DE DATOS
# ---------------------
conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

# Tablas
cursor.executescript(r"""
CREATE TABLE IF NOT EXISTS clientes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nombre TEXT UNIQUE,
    telefono TEXT,
    direccion TEXT,
    email TEXT
);

CREATE TABLE IF NOT EXISTS proveedores (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nombre TEXT,
    telefono TEXT,
    contacto TEXT
);

CREATE TABLE IF NOT EXISTS stock (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    producto TEXT UNIQUE,
    cantidad INTEGER,
    proveedor_id INTEGER,
    FOREIGN KEY(proveedor_id) REFERENCES proveedores(id)
);

CREATE TABLE IF NOT EXISTS presupuestos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    cliente_id INTEGER,
    fecha TEXT,
    detalle TEXT,
    monto REAL,
    FOREIGN KEY(cliente_id) REFERENCES clientes(id)
);

CREATE TABLE IF NOT EXISTS presupuesto_detalle (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    presupuesto_id INTEGER,
    producto_id INTEGER,
    cantidad INTEGER,
    FOREIGN KEY(presupuesto_id) REFERENCES presupuestos(id),
    FOREIGN KEY(producto_id) REFERENCES stock(id)
);

CREATE TABLE IF NOT EXISTS movimientos (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tipo TEXT, -- 'INGRESO' | 'EGRESO'
    descripcion TEXT,
    monto REAL,
    fecha TEXT
);

CREATE TABLE IF NOT EXISTS configuracion (
    clave TEXT PRIMARY KEY,
    valor TEXT
);
""")
conn.commit()

# --- MIGRACION SIMPLE: Asegurar columnas en clientes ---
try:
    cursor.execute("PRAGMA table_info(clientes)")
    columnas_clientes = [info[1] for info in cursor.fetchall()]
    if 'telefono' not in columnas_clientes:
        cursor.execute("ALTER TABLE clientes ADD COLUMN telefono TEXT")
    if 'direccion' not in columnas_clientes:
        cursor.execute("ALTER TABLE clientes ADD COLUMN direccion TEXT")
    if 'email' not in columnas_clientes:
        cursor.execute("ALTER TABLE clientes ADD COLUMN email TEXT")
    conn.commit()
except sqlite3.Error:
    # La tabla clientes podria no existir aun, lo cual es manejado por executescript.
    # No hacemos nada aqui.
    pass
# --- FIN MIGRACION ---

# --- MIGRACION SIMPLE: Asegurar columna en stock ---
try:
    cursor.execute("PRAGMA table_info(stock)")
    columnas_stock = [info[1] for info in cursor.fetchall()]
    if 'proveedor_id' not in columnas_stock:
        cursor.execute("ALTER TABLE stock ADD COLUMN proveedor_id INTEGER REFERENCES proveedores(id)")
    conn.commit()
except sqlite3.Error:
    # La tabla stock podria no existir aun.
    pass
# --- FIN MIGRACION ---

# --- MIGRACION PRESUPUESTOS ---
try:
    cursor.execute("PRAGMA table_info(clientes)")
    columnas_clientes = [info[1] for info in cursor.fetchall()]
    if 'dni_cuit' not in columnas_clientes:
        cursor.execute("ALTER TABLE clientes ADD COLUMN dni_cuit TEXT")

    cursor.execute("PRAGMA table_info(presupuestos)")
    columnas_presupuestos = [info[1] for info in cursor.fetchall()]
    if 'obra_descripcion' not in columnas_presupuestos:
        cursor.execute("ALTER TABLE presupuestos ADD COLUMN obra_descripcion TEXT")
    if 'fecha_inicio' not in columnas_presupuestos:
        cursor.execute("ALTER TABLE presupuestos ADD COLUMN fecha_inicio TEXT")
    if 'condiciones' not in columnas_presupuestos:
        cursor.execute("ALTER TABLE presupuestos ADD COLUMN condiciones TEXT")
    if 'subtotal' not in columnas_presupuestos:
        cursor.execute("ALTER TABLE presupuestos ADD COLUMN subtotal REAL")
    if 'iva' not in columnas_presupuestos:
        cursor.execute("ALTER TABLE presupuestos ADD COLUMN iva REAL")
    if 'estado' not in columnas_presupuestos:
        cursor.execute("ALTER TABLE presupuestos ADD COLUMN estado TEXT DEFAULT 'Pendiente'")

    # La columna 'monto' ahora representara el 'total'
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS presupuesto_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        presupuesto_id INTEGER,
        concepto TEXT,
        unidad TEXT,
        cantidad REAL,
        precio_unitario REAL,
        subtotal REAL,
        FOREIGN KEY(presupuesto_id) REFERENCES presupuestos(id)
    )
    """)
    conn.commit()
except sqlite3.Error as e:
    print(f"Error en migración de presupuestos: {e}")
    pass
# --- FIN MIGRACION ---


# ---------------------
# GESTION DE CONFIGURACION
# ---------------------
def obtener_config(clave):
    cursor.execute("SELECT valor FROM configuracion WHERE clave = ?", (clave,))
    resultado = cursor.fetchone()
    return resultado[0] if resultado else ""

def guardar_config(clave, valor):
    cursor.execute("REPLACE INTO configuracion (clave, valor) VALUES (?, ?)", (clave, valor))
    conn.commit()

# --- Inicializar configuración por defecto ---
default_config = {
    "EMPRESA_NOMBRE": "Stoplac",
    "EMPRESA_CUIT": "30-xxxxxxxx-x",
    "EMPRESA_DIRECCION": "Calle Falsa 123, Ciudad",
    "EMPRESA_TELEFONO": "11-1234-5678",
    "EMPRESA_CONTACTO": "info@stoplac.com",
    "EMPRESA_LOGO": "logo.png",
    "SMTP_SERVER": "",
    "SMTP_PORT": "",
    "SMTP_USER": "",
    "SMTP_PASS": ""
}
for clave, valor in default_config.items():
    cursor.execute("INSERT OR IGNORE INTO configuracion (clave, valor) VALUES (?, ?)", (clave, valor))
conn.commit()

LOGO_PATH = obtener_config("EMPRESA_LOGO")
if not LOGO_PATH:
    LOGO_PATH = "logo.png" # Fallback


# Inicializar stock / proveedores por defecto (solo si vacio)
cursor.execute("SELECT COUNT(*) FROM stock")
if cursor.fetchone()[0] == 0:
    # crear ejemplo de proveedores
    cursor.execute("INSERT OR IGNORE INTO proveedores(nombre, telefono, contacto) VALUES (?,?,?)", ("Proveedor A", "", ""))
    cursor.execute("INSERT OR IGNORE INTO proveedores(nombre, telefono, contacto) VALUES (?,?,?)", ("Proveedor B", "", ""))
    conn.commit()
    cursor.execute("SELECT id FROM proveedores LIMIT 1")
    prov_id = cursor.fetchone()[0]
    sample = {
        "Placa Standard": 50,
        "Placa Anti-humedad": 30,
        "Placa Decorativa": 20,
        "Pintura Blanca": 100,
        "Pintura Color": 80,
        "Barniz": 50
    }
    for p, q in sample.items():
        cursor.execute("INSERT OR IGNORE INTO stock(producto, cantidad, proveedor_id) VALUES(?,?,?)", (p, q, prov_id))
    conn.commit()

# ---------------------
# UTILIDADES
# ---------------------

def cargar_stock_dict():
    cursor.execute("SELECT id, producto, cantidad, proveedor_id FROM stock")
    return {row[0]: {"producto": row[1], "cantidad": row[2], "proveedor_id": row[3]} for row in cursor.fetchall()}


def obtener_clientes():
    cursor.execute("SELECT id, nombre, telefono, direccion, dni_cuit, email FROM clientes ORDER BY nombre")
    return cursor.fetchall()


def obtener_proveedores():
    cursor.execute("SELECT id, nombre, telefono, contacto FROM proveedores ORDER BY nombre")
    return cursor.fetchall()


def fecha_now():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def enviar_email(destinatario, asunto, cuerpo, adjunto_path=None):
    remitente_email = obtener_config("SMTP_USER")
    remitente_pass = obtener_config("SMTP_PASS")
    smtp_server = obtener_config("SMTP_SERVER")
    smtp_port = obtener_config("SMTP_PORT")

    if not all([remitente_email, remitente_pass, smtp_server, smtp_port]):
        messagebox.showerror("Configuración Incompleta", "Faltan datos de configuración de SMTP en la pestaña 'Configuración'.")
        return False

    msg = MIMEMultipart()
    msg['From'] = remitente_email
    msg['To'] = destinatario
    msg['Subject'] = asunto
    msg.attach(MIMEText(cuerpo, 'plain'))

    if adjunto_path:
        try:
            with open(adjunto_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(adjunto_path)}")
            msg.attach(part)
        except Exception as e:
            messagebox.showerror("Error adjunto", f"No se pudo adjuntar el archivo: {e}")
            return False

    try:
        smtp_port_int = int(smtp_port)
        if smtp_port_int == 465:
            # Usar SSL desde el inicio
            server = smtplib.SMTP_SSL(smtp_server, smtp_port_int)
        else:
            # Usar STARTTLS
            server = smtplib.SMTP(smtp_server, smtp_port_int)
            server.starttls()
        
        server.login(remitente_email, remitente_pass)
        text = msg.as_string()
        server.sendmail(remitente_email, destinatario, text)
        server.quit()
        messagebox.showinfo("Éxito", "Email enviado correctamente.")
        return True
    except smtplib.SMTPAuthenticationError:
        messagebox.showerror("Error de Autenticación", 
            "Usuario o contraseña incorrectos.\n\n"
            "Si usa Gmail, es posible que necesite una 'Contraseña de Aplicación' "
            "en lugar de su contraseña habitual. Puede generarla en la configuración "
            "de seguridad de su cuenta de Google.")
        return False
    except Exception as e:
        messagebox.showerror("Error de envío", f"No se pudo enviar el email: {e}")
        return False

# ---------------------
# PDF GENERADORES
# ---------------------

def draw_logo_on_canvas(c, x, y, width=100):
    logo_path_from_db = obtener_config("EMPRESA_LOGO")
    if not logo_path_from_db or not os.path.exists(logo_path_from_db):
        return 0  # No hay logo o no se encuentra

    try:
        logo = ImageReader(logo_path_from_db)
        logo_width, logo_height = logo.getSize()
        aspect = logo_height / float(logo_width)
        logo_height = width * aspect
        c.drawImage(logo, x, y - logo_height, width=width, height=logo_height, preserveAspectRatio=True, mask='auto')
        return logo_height
    except Exception:
        return 0


def generar_pdf_presupuesto(presupuesto_id, path=None):
    cursor.execute("SELECT p.id, c.nombre, p.fecha, p.detalle, p.monto FROM presupuestos p JOIN clientes c ON p.cliente_id=c.id WHERE p.id=?", (presupuesto_id,))
    row = cursor.fetchone()
    if not row:
        raise ValueError("Presupuesto no encontrado")
    _, cliente, fecha_pres, detalle, monto = row
    cursor.execute("SELECT pd.producto_id, pd.cantidad, s.producto FROM presupuesto_detalle pd JOIN stock s ON pd.producto_id=s.id WHERE pd.presupuesto_id=?", (presupuesto_id,))
    productos = [(r[2], r[1]) for r in cursor.fetchall()]

    if not path:
        path = os.path.join(PDF_DIR, f"Presupuesto_{presupuesto_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

    c = canvas.Canvas(path, pagesize=A4)
    width, height = A4
    draw_logo_on_canvas(c)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(160, height - 80, "Presupuesto - Control Premium")
    c.setFont("Helvetica", 12)
    c.drawString(40, height - 140, f"ID: {presupuesto_id}")
    c.drawString(40, height - 160, f"Cliente: {cliente}")
    c.drawString(40, height - 180, f"Fecha registro: {fecha_pres}")
    c.drawString(40, height - 200, "Detalle del trabajo:")
    y = height - 220
    for line in (detalle or '').splitlines():
        c.drawString(60, y, line)
        y -= 14
    y -= 6
    c.drawString(40, y, "Productos utilizados:")
    y -= 18
    for prod, qty in productos:
        c.drawString(60, y, f"- {prod}: {qty}")
        y -= 14
    y -= 6
    c.drawString(40, y, f"Monto total: ${monto:.2f}")
    c.showPage()
    c.save()
    return path


def generar_pdf_reporte_balance(path=None):
    cursor.execute("SELECT tipo, descripcion, monto, fecha FROM movimientos ORDER BY fecha")
    rows = cursor.fetchall()
    if not path:
        path = os.path.join(REPORT_DIR, f"Balance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    c = canvas.Canvas(path, pagesize=A4)
    width, height = A4
    draw_logo_on_canvas(c)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(160, height - 80, "Control Premium - Balance de movimientos")
    c.setFont("Helvetica", 12)
    y = height - 120
    saldo = 0.0
    for tipo, desc, monto, fecha in rows:
        sign = 1 if tipo == 'INGRESO' else -1
        saldo += sign * monto
        c.drawString(40, y, f"{fecha} | {tipo} | ${monto:.2f} | {desc}")
        y -= 14
        if y < 60:
            c.showPage()
            y = height - 60
    c.drawString(40, y - 20, f"Saldo actual: ${saldo:.2f}")
    c.showPage()
    c.save()
    return path

# ---------------------
# INTERFAZ
# ---------------------
root = tk.Tk()
root.title(APP_TITLE)
root.geometry('1100x700')

# --- ESTILO PROFESIONAL ---
BG_COLOR = "#ECECEC"  # Gris claro
TEXT_COLOR = "#333333"
PRIMARY_COLOR = "#0078D7"  # Azul
SUCCESS_COLOR = "#28A745"  # Verde
DANGER_COLOR = "#DC3545"  # Rojo
INFO_COLOR = "#607D8B"     # Gris azulado
WARNING_COLOR = "#FF9800"  # Naranja

FONT_FAMILY = "Calibri"
FONT_NORMAL = (FONT_FAMILY, 11)
FONT_BOLD = (FONT_FAMILY, 11, "bold")
FONT_TITLE = (FONT_FAMILY, 18, "bold")
FONT_H2 = (FONT_FAMILY, 14, "bold")

root.configure(bg=BG_COLOR)

# Cargar logo (se muestra arriba de notebook)
logo_img_tk = None
if os.path.exists(LOGO_PATH):
    try:
        img = Image.open(LOGO_PATH)
        img = img.resize((120, 120), Image.LANCZOS)
        logo_img_tk = ImageTk.PhotoImage(img)
    except Exception:
        logo_img_tk = None

# Marco superior con logo y titulo
top_frame = tk.Frame(root, bg=BG_COLOR)
top_frame.pack(fill='x', padx=10, pady=5)
if logo_img_tk:
    lbl_logo_top = tk.Label(top_frame, image=logo_img_tk, bg=BG_COLOR)
    lbl_logo_top.pack(side='left', padx=10)
else:
    lbl_logo_top = tk.Label(top_frame, text='Control Premium', font=(FONT_FAMILY, 24, 'bold'), bg=BG_COLOR, fg=PRIMARY_COLOR)
    lbl_logo_top.pack(side='left', padx=10)

lbl_title = tk.Label(top_frame, text='Gestión Profesional', font=FONT_TITLE, bg=BG_COLOR, fg=TEXT_COLOR)
lbl_title.pack(side='left', padx=10)

# Notebook
notebook = ttk.Notebook(root)
notebook.pack(fill='both', expand=True, padx=10, pady=10)

# Estilo para widgets ttk
style = ttk.Style()
style.theme_use('clam')

style.configure("TNotebook", background=BG_COLOR, borderwidth=0)
style.configure("TNotebook.Tab",
                background="#D0D0D0",
                foreground=TEXT_COLOR,
                padding=[12, 8],
                font=FONT_BOLD)
style.map("TNotebook.Tab",
          background=[("selected", PRIMARY_COLOR)],
          foreground=[("selected", "white")])

style.configure("TFrame", background=BG_COLOR)

style.configure("Treeview",
                background="white",
                foreground=TEXT_COLOR,
                rowheight=28,
                fieldbackground="white",
                font=FONT_NORMAL)
style.map('Treeview', background=[('selected', PRIMARY_COLOR)], foreground=[('selected', 'white')])

style.configure("Treeview.Heading",
                background="#DFDFDF",
                foreground=TEXT_COLOR,
                font=FONT_BOLD,
                padding=[6, 6])

# ---------- TAB CLIENTES ----------
frame_clientes = ttk.Frame(notebook, style="TFrame")
notebook.add(frame_clientes, text='Clientes')

# Contenido clientes
clientes_top = tk.Frame(frame_clientes, bg=BG_COLOR)
clientes_top.pack(fill='x', pady=6)

lbl_c = tk.Label(clientes_top, text='Gestion de Clientes', font=FONT_H2, bg=BG_COLOR, fg=TEXT_COLOR)
lbl_c.pack(side='left', padx=6)

# Form cliente
frm_c_form = tk.Frame(frame_clientes, bg=BG_COLOR)
frm_c_form.pack(fill='x', padx=8, pady=6)

tk.Label(frm_c_form, text='Nombre:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=0, column=0, sticky='w')
entry_c_nombre = tk.Entry(frm_c_form, width=40, font=FONT_NORMAL, relief='solid', bd=1)
entry_c_nombre.grid(row=0, column=1, padx=6, pady=2)

tk.Label(frm_c_form, text='DNI/CUIT:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=1, column=0, sticky='w')
entry_c_dnicuit = tk.Entry(frm_c_form, width=25, font=FONT_NORMAL, relief='solid', bd=1)
entry_c_dnicuit.grid(row=1, column=1, sticky='w', padx=6, pady=2)

tk.Label(frm_c_form, text='Teléfono:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=2, column=0, sticky='w')
entry_c_telefono = tk.Entry(frm_c_form, width=25, font=FONT_NORMAL, relief='solid', bd=1)
entry_c_telefono.grid(row=2, column=1, sticky='w', padx=6, pady=2)

tk.Label(frm_c_form, text='Dirección:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=3, column=0, sticky='w')
entry_c_direccion = tk.Entry(frm_c_form, width=60, font=FONT_NORMAL, relief='solid', bd=1)
entry_c_direccion.grid(row=3, column=1, padx=6, pady=2)

tk.Label(frm_c_form, text='Email:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=4, column=0, sticky='w')
entry_c_email = tk.Entry(frm_c_form, width=60, font=FONT_NORMAL, relief='solid', bd=1)
entry_c_email.grid(row=4, column=1, padx=6, pady=2)

frm_c_btn = tk.Frame(frm_c_form, bg=BG_COLOR)
frm_c_btn.grid(row=0, column=2, rowspan=5, padx=10)


def agregar_cliente():
    nombre = entry_c_nombre.get().strip()
    tel = entry_c_telefono.get().strip()
    dire = entry_c_direccion.get().strip()
    dnicuit = entry_c_dnicuit.get().strip()
    email = entry_c_email.get().strip()
    if not nombre:
        messagebox.showwarning('Falta dato', 'El nombre es obligatorio')
        return
    cursor.execute('INSERT OR IGNORE INTO clientes(nombre, telefono, direccion, dni_cuit, email) VALUES(?,?,?,?,?)', (nombre, tel, dire, dnicuit, email))
    conn.commit()
    entry_c_nombre.delete(0, 'end')
    entry_c_telefono.delete(0, 'end')
    entry_c_direccion.delete(0, 'end')
    entry_c_dnicuit.delete(0, 'end')
    entry_c_email.delete(0, 'end')
    actualizar_tree_clientes()


def editar_cliente():
    sel = tree_clientes.selection()
    if not sel:
        messagebox.showwarning('Seleccionar', 'Selecciona un cliente')
        return
    item = tree_clientes.item(sel[0], 'values')
    cid = item[0]
    nuevo_nombre = simpledialog.askstring('Editar', 'Nombre:', initialvalue=item[1])
    if not nuevo_nombre:
        return
    nuevo_dnicuit = simpledialog.askstring('Editar', 'DNI/CUIT:', initialvalue=item[4] or '')
    nuevo_tel = simpledialog.askstring('Editar', 'Teléfono:', initialvalue=item[2] or '')
    nuevo_dir = simpledialog.askstring('Editar', 'Dirección:', initialvalue=item[3] or '')
    nuevo_email = simpledialog.askstring('Editar', 'Email:', initialvalue=item[5] or '')
    cursor.execute('UPDATE clientes SET nombre=?, telefono=?, direccion=?, dni_cuit=?, email=? WHERE id=?', (nuevo_nombre, nuevo_tel or '', nuevo_dir or '', nuevo_dnicuit or '', nuevo_email or '', cid))
    conn.commit()
    actualizar_tree_clientes()


def eliminar_cliente():
    sel = tree_clientes.selection()
    if not sel:
        messagebox.showwarning('Seleccionar', 'Selecciona un cliente')
        return
    item = tree_clientes.item(sel[0], 'values')
    cid = item[0]
    if messagebox.askyesno('Confirmar', f"Eliminar cliente {item[1]}? Esto no eliminará presupuestos existentes." ):
        cursor.execute('DELETE FROM clientes WHERE id=?', (cid,))
        conn.commit()
        actualizar_tree_clientes()

btn_c_add = tk.Button(frm_c_btn, text='Agregar', command=agregar_cliente, bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_c_add.pack(fill='x', pady=3)
btn_c_edit = tk.Button(frm_c_btn, text='Editar', command=editar_cliente, bg=PRIMARY_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_c_edit.pack(fill='x', pady=3)
btn_c_del = tk.Button(frm_c_btn, text='Eliminar', command=eliminar_cliente, bg=DANGER_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_c_del.pack(fill='x', pady=3)

# Treeview clientes
cols = ('id', 'nombre', 'telefono', 'direccion', 'dni_cuit', 'email')
tree_clientes = ttk.Treeview(frame_clientes, columns=cols, show='headings', height=8)
for c in cols:
    tree_clientes.heading(c, text=c.capitalize())
tree_clientes.pack(fill='x', padx=8, pady=6)


def actualizar_tree_clientes():
    for r in tree_clientes.get_children():
        tree_clientes.delete(r)
    
    clientes = obtener_clientes()
    for row in clientes:
        tree_clientes.insert('', 'end', values=row)
    
    # Actualizar combobox en presupuestos
    try:
        clientes_nombres = [c[1] for c in clientes]
        combo_pres_cliente_seleccionar['values'] = clientes_nombres
    except (tk.TclError, NameError):
        # El widget puede no existir aun al inicio
        pass

actualizar_tree_clientes()

# ---------- TAB PROVEEDORES ----------
frame_proveedores = ttk.Frame(notebook, style="TFrame")
notebook.add(frame_proveedores, text='Proveedores')

frm_p_top = tk.Frame(frame_proveedores, bg=BG_COLOR)
frm_p_top.pack(fill='x', pady=6)

lbl_p = tk.Label(frm_p_top, text='Gestion de Proveedores', font=FONT_H2, bg=BG_COLOR, fg=TEXT_COLOR)
lbl_p.pack(side='left', padx=6)

frm_p_form = tk.Frame(frame_proveedores, bg=BG_COLOR)
frm_p_form.pack(fill='x', padx=8, pady=6)

tk.Label(frm_p_form, text='Nombre:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=0, column=0, sticky='w')
entry_p_nombre = tk.Entry(frm_p_form, width=40, font=FONT_NORMAL, relief='solid', bd=1)
entry_p_nombre.grid(row=0, column=1, padx=6, pady=2)

tk.Label(frm_p_form, text='Teléfono:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=1, column=0, sticky='w')
entry_p_telefono = tk.Entry(frm_p_form, width=25, font=FONT_NORMAL, relief='solid', bd=1)
entry_p_telefono.grid(row=1, column=1, sticky='w', padx=6, pady=2)

tk.Label(frm_p_form, text='Contacto:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=2, column=0, sticky='w')
entry_p_contacto = tk.Entry(frm_p_form, width=60, font=FONT_NORMAL, relief='solid', bd=1)
entry_p_contacto.grid(row=2, column=1, padx=6, pady=2)

frm_p_btn = tk.Frame(frm_p_form, bg=BG_COLOR)
frm_p_btn.grid(row=0, column=2, rowspan=3, padx=10)


def agregar_proveedor():
    nombre = entry_p_nombre.get().strip()
    tel = entry_p_telefono.get().strip()
    cont = entry_p_contacto.get().strip()
    if not nombre:
        messagebox.showwarning('Falta dato', 'El nombre es obligatorio')
        return
    cursor.execute('INSERT INTO proveedores(nombre, telefono, contacto) VALUES(?,?,?)', (nombre, tel, cont))
    conn.commit()
    entry_p_nombre.delete(0, 'end')
    entry_p_telefono.delete(0, 'end')
    entry_p_contacto.delete(0, 'end')
    actualizar_tree_proveedores()


def editar_proveedor():
    sel = tree_proveedores.selection()
    if not sel:
        messagebox.showwarning('Seleccionar', 'Selecciona un proveedor')
        return
    item = tree_proveedores.item(sel[0], 'values')
    pid = item[0]
    nuevo_nombre = simpledialog.askstring('Editar', 'Nombre:', initialvalue=item[1])
    if not nuevo_nombre:
        return
    nuevo_tel = simpledialog.askstring('Editar', 'Teléfono:', initialvalue=item[2] or '')
    nuevo_cont = simpledialog.askstring('Editar', 'Contacto:', initialvalue=item[3] or '')
    cursor.execute('UPDATE proveedores SET nombre=?, telefono=?, contacto=? WHERE id=?', (nuevo_nombre, nuevo_tel or '', nuevo_cont or '', pid))
    conn.commit()
    actualizar_tree_proveedores()


def eliminar_proveedor():
    sel = tree_proveedores.selection()
    if not sel:
        messagebox.showwarning('Seleccionar', 'Selecciona un proveedor')
        return
    item = tree_proveedores.item(sel[0], 'values')
    pid = item[0]
    if messagebox.askyesno('Confirmar', f"Eliminar proveedor {item[1]}? Esto no eliminará productos asociados automáticamente." ):
        cursor.execute('DELETE FROM proveedores WHERE id=?', (pid,))
        conn.commit()
        actualizar_tree_proveedores()

btn_p_add = tk.Button(frm_p_btn, text='Agregar', command=agregar_proveedor, bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_p_add.pack(fill='x', pady=3)
btn_p_edit = tk.Button(frm_p_btn, text='Editar', command=editar_proveedor, bg=PRIMARY_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_p_edit.pack(fill='x', pady=3)
btn_p_del = tk.Button(frm_p_btn, text='Eliminar', command=eliminar_proveedor, bg=DANGER_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_p_del.pack(fill='x', pady=3)

cols_p = ('id', 'nombre', 'telefono', 'contacto')
tree_proveedores = ttk.Treeview(frame_proveedores, columns=cols_p, show='headings', height=8)
for c in cols_p:
    tree_proveedores.heading(c, text=c.capitalize())
tree_proveedores.pack(fill='x', padx=8, pady=6)


def actualizar_tree_proveedores():
    for r in tree_proveedores.get_children():
        tree_proveedores.delete(r)
    for row in obtener_proveedores():
        tree_proveedores.insert('', 'end', values=row)

actualizar_tree_proveedores()

# ---------- TAB STOCK ----------
frame_stock = ttk.Frame(notebook, style="TFrame")
notebook.add(frame_stock, text='Stock')

frm_s_top = tk.Frame(frame_stock, bg=BG_COLOR)
frm_s_top.pack(fill='x', pady=6)

lbl_s = tk.Label(frm_s_top, text='Inventario', font=FONT_H2, bg=BG_COLOR, fg=TEXT_COLOR)
lbl_s.pack(side='left', padx=6)

entry_s_buscar = tk.Entry(frm_s_top, width=30, font=FONT_NORMAL, relief='solid', bd=1)
entry_s_buscar.pack(side='left', padx=6)
btn_s_buscar = tk.Button(frm_s_top, text='Buscar', command=lambda: buscar_stock(entry_s_buscar.get()), bg=INFO_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_s_buscar.pack(side='left', padx=6)

def exportar_stock_excel():
    """Exporta el stock actual a un archivo Excel."""
    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
        title="Guardar stock como..."
    )
    if not filepath:
        return

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Stock"

        # Escribir encabezados
        headers = ["ID", "Producto", "Cantidad", "ID Proveedor"]
        sheet.append(headers)

        # Obtener y escribir datos
        cursor.execute('SELECT id, producto, cantidad, proveedor_id FROM stock ORDER BY producto')
        for row in cursor.fetchall():
            sheet.append(row)

        workbook.save(filepath)
        messagebox.showinfo("Éxito", f"Stock exportado correctamente a\n{filepath}")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar el stock: {e}")

def importar_stock_excel():
    """Importa productos desde un archivo Excel, actualizando o agregando según corresponda."""
    filepath = filedialog.askopenfilename(
        filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
        title="Importar stock desde..."
    )
    if not filepath:
        return

    if not messagebox.askyesno("Confirmar Importación", 
        "Esto actualizará el stock con los datos del archivo. "
        "Las cantidades de productos existentes se sumarán. "
        "¿Desea continuar?"):
        return

    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        # Omitir encabezado
        iter_rows = iter(sheet.rows)
        next(iter_rows)

        for row in iter_rows:
            # Asumiendo formato: Producto (col A), Cantidad (col B)
            producto = row[1].value
            cantidad = row[2].value
            
            if not producto or cantidad is None:
                continue # Ignorar filas vacías

            try:
                cantidad = int(cantidad)
            except (ValueError, TypeError):
                print(f"Omitiendo producto '{producto}' por cantidad inválida: '{cantidad}'")
                continue

            cursor.execute("SELECT id, cantidad FROM stock WHERE producto = ?", (producto,))
            existente = cursor.fetchone()

            if existente:
                # Si existe, sumar a la cantidad actual
                nueva_cantidad = existente[1] + cantidad
                cursor.execute("UPDATE stock SET cantidad = ? WHERE id = ?", (nueva_cantidad, existente[0]))
            else:
                # Si no existe, insertar nuevo (proveedor se deja nulo)
                cursor.execute('INSERT INTO stock(producto, cantidad) VALUES(?,?)', (producto, cantidad))
        
        conn.commit()
        actualizar_tree_stock()
        messagebox.showinfo("Éxito", "Stock importado y actualizado correctamente.")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo importar el stock: {e}")


btn_s_exportar = tk.Button(frm_s_top, text='Exportar a Excel', command=exportar_stock_excel, bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_s_exportar.pack(side='right', padx=6)
btn_s_importar = tk.Button(frm_s_top, text='Importar desde Excel', command=importar_stock_excel, bg=PRIMARY_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_s_importar.pack(side='right', padx=6)


# Form stock: producto, cantidad, proveedor
frm_s_form = tk.Frame(frame_stock, bg=BG_COLOR)
frm_s_form.pack(fill='x', padx=8, pady=6)

tk.Label(frm_s_form, text='Producto:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=0, column=0, sticky='w')
entry_s_producto = tk.Entry(frm_s_form, width=40, font=FONT_NORMAL, relief='solid', bd=1)
entry_s_producto.grid(row=0, column=1, padx=6, pady=2)

tk.Label(frm_s_form, text='Cantidad:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=1, column=0, sticky='w')
entry_s_cantidad = tk.Entry(frm_s_form, width=10, font=FONT_NORMAL, relief='solid', bd=1)
entry_s_cantidad.grid(row=1, column=1, sticky='w', padx=6, pady=2)

# proveedor combobox
proveedores = obtener_proveedores()
prov_dict = {str(p[0]): p[1] for p in proveedores}
prov_values = [f"{p[0]} - {p[1]}" for p in proveedores]
combo_s_prov = ttk.Combobox(frm_s_form, values=prov_values, width=30)
combo_s_prov.grid(row=2, column=1, sticky='w', padx=6, pady=2)

tk.Label(frm_s_form, text='Proveedor:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=2, column=0, sticky='w')
frm_s_btn = tk.Frame(frm_s_form, bg=BG_COLOR)
frm_s_btn.grid(row=0, column=2, rowspan=3, padx=10)


def agregar_stock():
    prod = entry_s_producto.get().strip()
    try:
        cant = int(entry_s_cantidad.get().strip())
    except Exception:
        messagebox.showwarning('Error', 'Cantidad debe ser un numero entero')
        return
    prov_sel = combo_s_prov.get()
    prov_id = None
    if prov_sel:
        prov_id = int(prov_sel.split(' - ')[0])
    if not prod:
        messagebox.showwarning('Falta dato', 'Producto es obligatorio')
        return

    # Verificar si el producto ya existe
    cursor.execute("SELECT id, cantidad FROM stock WHERE producto = ?", (prod,))
    existente = cursor.fetchone()

    if existente:
        # Si existe, actualizar la cantidad
        nueva_cantidad = existente[1] + cant
        cursor.execute("UPDATE stock SET cantidad = ?, proveedor_id = ? WHERE id = ?", (nueva_cantidad, prov_id, existente[0]))
    else:
        # Si no existe, insertar un nuevo registro
        cursor.execute('INSERT INTO stock(producto, cantidad, proveedor_id) VALUES(?,?,?)', (prod, cant, prov_id))
    
    conn.commit()
    entry_s_producto.delete(0, 'end')
    entry_s_cantidad.delete(0, 'end')
    actualizar_tree_stock()


def editar_stock():
    sel = tree_stock_view.selection()
    if not sel:
        messagebox.showwarning('Seleccionar', 'Selecciona un item de stock')
        return
    item = tree_stock_view.item(sel[0], 'values')
    sid = item[0]
    nuevo_prod = simpledialog.askstring('Editar', 'Producto:', initialvalue=item[1])
    nuevo_cant = simpledialog.askinteger('Editar', 'Cantidad:', initialvalue=int(item[2] or 0))
    if nuevo_prod is None or nuevo_cant is None:
        return
    cursor.execute('UPDATE stock SET producto=?, cantidad=? WHERE id=?', (nuevo_prod, nuevo_cant, sid))
    conn.commit()
    actualizar_tree_stock()


def eliminar_stock():
    sel = tree_stock_view.selection()
    if not sel:
        messagebox.showwarning('Seleccionar', 'Selecciona un item de stock')
        return
    item = tree_stock_view.item(sel[0], 'values')
    sid = item[0]
    if messagebox.askyesno('Confirmar', f"Eliminar producto {item[1]}?"):
        cursor.execute('DELETE FROM stock WHERE id=?', (sid,))
        conn.commit()
        actualizar_tree_stock()


def reponer_stock():
    sel = tree_stock_view.selection()
    if not sel:
        messagebox.showwarning('Seleccionar', 'Selecciona un item de stock para reponer')
        return
    item = tree_stock_view.item(sel[0], 'values')
    sid = item[0]
    producto_nombre = item[1]
    cantidad_actual = int(item[2])

    cantidad_a_reponer = simpledialog.askinteger('Reponer Stock', f'Ingrese la cantidad a reponer para "{producto_nombre}":', initialvalue=0)

    if cantidad_a_reponer is not None and cantidad_a_reponer > 0:
        nueva_cantidad = cantidad_actual + cantidad_a_reponer
        cursor.execute('UPDATE stock SET cantidad=? WHERE id=?', (nueva_cantidad, sid))
        conn.commit()
        actualizar_tree_stock()
    elif cantidad_a_reponer is not None:
        messagebox.showwarning('Cantidad inválida', 'La cantidad a reponer debe ser mayor que cero.')


def buscar_stock(termino):
    for r in tree_stock_view.get_children():
        tree_stock_view.delete(r)
    
    termino = termino.strip()
    if not termino:
        # Si no hay término de búsqueda, mostrar todo
        cursor.execute('SELECT id, producto, cantidad, proveedor_id FROM stock ORDER BY producto')
    else:
        # Buscar productos que contengan el término
        cursor.execute('SELECT id, producto, cantidad, proveedor_id FROM stock WHERE producto LIKE ? ORDER BY producto', ('%' + termino + '%',))
    
    for row in cursor.fetchall():
        tree_stock_view.insert('', 'end', values=row)

btn_s_add = tk.Button(frm_s_btn, text='Agregar', command=agregar_stock, bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_s_add.pack(fill='x', pady=3)
btn_s_edit = tk.Button(frm_s_btn, text='Editar', command=editar_stock, bg=PRIMARY_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_s_edit.pack(fill='x', pady=3)
btn_s_reponer = tk.Button(frm_s_btn, text='Reponer', command=reponer_stock, bg=WARNING_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_s_reponer.pack(fill='x', pady=3)
btn_s_del = tk.Button(frm_s_btn, text='Eliminar', command=eliminar_stock, bg=DANGER_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_s_del.pack(fill='x', pady=3)

cols_s = ('id', 'producto', 'cantidad', 'proveedor_id')
tree_stock_view = ttk.Treeview(frame_stock, columns=cols_s, show='headings', height=8)
for c in cols_s:
    tree_stock_view.heading(c, text=c.capitalize())
tree_stock_view.pack(fill='x', padx=8, pady=6)


def actualizar_tree_stock():
    for r in tree_stock_view.get_children():
        tree_stock_view.delete(r)
    cursor.execute('SELECT id, producto, cantidad, proveedor_id FROM stock ORDER BY producto')
    for row in cursor.fetchall():
        tree_stock_view.insert('', 'end', values=row)

actualizar_tree_stock()

# ---------- TAB PRESUPUESTOS ----------
frame_pres = ttk.Frame(notebook, style="TFrame")
notebook.add(frame_pres, text='Presupuestos')

# --- Contenedor principal ---
pres_main_frame = tk.Frame(frame_pres, bg=BG_COLOR)
pres_main_frame.pack(fill='both', expand=True)
pres_main_frame.columnconfigure(0, weight=1)
pres_main_frame.rowconfigure(1, weight=1) # Permite que el frame de items se expanda

# --- Lógica ---
def obtener_detalle_cliente_por_nombre(nombre):
    cursor.execute("SELECT nombre, dni_cuit, direccion, telefono FROM clientes WHERE nombre = ?", (nombre,))
    return cursor.fetchone()

def autocompletar_cliente(event):
    nombre_seleccionado = combo_pres_cliente_seleccionar.get()
    if not nombre_seleccionado:
        return

    cliente_data = obtener_detalle_cliente_por_nombre(nombre_seleccionado)
    
    if cliente_data:
        nombre, dni_cuit, domicilio, telefono = cliente_data
        
        # Limpiar campos
        entry_pres_cliente_nombre.delete(0, 'end')
        entry_pres_cliente_cuit.delete(0, 'end')
        entry_pres_cliente_domicilio.delete(0, 'end')
        entry_pres_cliente_telefono.delete(0, 'end')
        
        # Llenar campos
        entry_pres_cliente_nombre.insert(0, nombre or "")
        entry_pres_cliente_cuit.insert(0, dni_cuit or "")
        entry_pres_cliente_domicilio.insert(0, domicilio or "")
        entry_pres_cliente_telefono.insert(0, telefono or "")

def calcular_totales_presupuesto():
    subtotal = 0.0
    for item_id in tree_items.get_children():
        valores = tree_items.item(item_id, 'values')
        try:
            item_subtotal = float(valores[4])
            subtotal += item_subtotal
        except (ValueError, IndexError):
            continue # Ignorar si el subtotal no es un número válido

    if iva_var.get():
        iva = subtotal * 0.21
    else:
        iva = 0.0
        
    total = subtotal + iva

    # Habilitar edición para actualizar
    entry_subtotal.config(state='normal')
    entry_iva.config(state='normal')
    entry_total.config(state='normal')

    entry_subtotal.delete(0, 'end')
    entry_subtotal.insert(0, f"{subtotal:.2f}")
    entry_iva.delete(0, 'end')
    entry_iva.insert(0, f"{iva:.2f}")
    entry_total.delete(0, 'end')
    entry_total.insert(0, f"{total:.2f}")

    # Volver a solo lectura
    entry_subtotal.config(state='readonly')
    entry_iva.config(state='readonly')
    entry_total.config(state='readonly')


def agregar_item_presupuesto():
    # --- Ventana Toplevel para agregar/editar item ---
    dialog = tk.Toplevel(root)
    dialog.title("Agregar Ítem al Presupuesto")
    dialog.geometry("400x250")
    dialog.resizable(False, False)
    dialog.transient(root)
    dialog.grab_set()
    dialog.configure(bg=BG_COLOR)

    frm_dialog = tk.Frame(dialog, bg=BG_COLOR, padx=15, pady=15)
    frm_dialog.pack(fill='both', expand=True)

    # --- Campos del formulario ---
    tk.Label(frm_dialog, text="Concepto:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=0, column=0, sticky='w', pady=5)
    
    # Combobox con productos de stock
    stock_items = [info['producto'] for info in cargar_stock_dict().values()]
    combo_concepto = ttk.Combobox(frm_dialog, values=stock_items, width=38, font=FONT_NORMAL)
    combo_concepto.grid(row=0, column=1, sticky='w', padx=5)

    tk.Label(frm_dialog, text="Unidad:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=1, column=0, sticky='w', pady=5)
    entry_unidad = tk.Entry(frm_dialog, width=15, font=FONT_NORMAL, relief='solid', bd=1)
    entry_unidad.grid(row=1, column=1, sticky='w', padx=5)
    entry_unidad.insert(0, "u")

    tk.Label(frm_dialog, text="Cantidad:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=2, column=0, sticky='w', pady=5)
    entry_cantidad = tk.Entry(frm_dialog, width=15, font=FONT_NORMAL, relief='solid', bd=1)
    entry_cantidad.grid(row=2, column=1, sticky='w', padx=5)

    tk.Label(frm_dialog, text="Precio Unitario:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=3, column=0, sticky='w', pady=5)
    entry_precio = tk.Entry(frm_dialog, width=15, font=FONT_NORMAL, relief='solid', bd=1)
    entry_precio.grid(row=3, column=1, sticky='w', padx=5)

    # --- Variable para almacenar resultado ---
    result = {}

    def on_ok():
        try:
            concepto = combo_concepto.get().strip()
            unidad = entry_unidad.get().strip()
            cantidad = float(entry_cantidad.get().strip())
            precio = float(entry_precio.get().strip())

            if not concepto or not unidad or cantidad <= 0 or precio < 0:
                messagebox.showwarning("Datos inválidos", "Por favor, complete todos los campos correctamente.", parent=dialog)
                return
            
            result['concepto'] = concepto
            result['unidad'] = unidad
            result['cantidad'] = cantidad
            result['precio'] = precio
            dialog.destroy()

        except ValueError:
            messagebox.showerror("Error de formato", "Cantidad y Precio deben ser números válidos.", parent=dialog)

    def on_cancel():
        dialog.destroy()

    # --- Botones ---
    frm_buttons = tk.Frame(frm_dialog, bg=BG_COLOR)
    frm_buttons.grid(row=4, column=0, columnspan=2, pady=15)

    btn_ok = tk.Button(frm_buttons, text="Aceptar", command=on_ok, bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
    btn_ok.pack(side='left', padx=10)
    btn_cancel = tk.Button(frm_buttons, text="Cancelar", command=on_cancel, bg=DANGER_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
    btn_cancel.pack(side='left', padx=10)

    # --- Esperar a que la ventana se cierre ---
    root.wait_window(dialog)

    if result:
        subtotal_item = result['cantidad'] * result['precio']
        tree_items.insert('', 'end', values=(
            result['concepto'], 
            result['unidad'], 
            result['cantidad'], 
            f"{result['precio']:.2f}", 
            f"{subtotal_item:.2f}"
        ))
        calcular_totales_presupuesto()


def modificar_item_presupuesto():
    sel = tree_items.selection()
    if not sel:
        messagebox.showwarning('Seleccionar', 'Selecciona un ítem para modificar.')
        return
    
    item_id = sel[0]
    valores = tree_items.item(item_id, 'values')

    # --- Ventana Toplevel para agregar/editar item ---
    dialog = tk.Toplevel(root)
    dialog.title("Modificar Ítem del Presupuesto")
    dialog.geometry("500x350")
    dialog.resizable(False, False)
    dialog.transient(root)
    dialog.grab_set()
    dialog.configure(bg=BG_COLOR)

    frm_dialog = tk.Frame(dialog, bg=BG_COLOR, padx=15, pady=15)
    frm_dialog.pack(fill='both', expand=True)

    # --- Campos del formulario ---
    tk.Label(frm_dialog, text="Concepto:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=0, column=0, sticky='w', pady=5)
    stock_items = [info['producto'] for info in cargar_stock_dict().values()]
    combo_concepto = ttk.Combobox(frm_dialog, values=stock_items, width=38, font=FONT_NORMAL)
    combo_concepto.grid(row=0, column=1, sticky='w', padx=5)
    combo_concepto.set(valores[0])

    tk.Label(frm_dialog, text="Unidad:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=1, column=0, sticky='w', pady=5)
    entry_unidad = tk.Entry(frm_dialog, width=15, font=FONT_NORMAL, relief='solid', bd=1)
    entry_unidad.grid(row=1, column=1, sticky='w', padx=5)
    entry_unidad.insert(0, valores[1])

    tk.Label(frm_dialog, text="Cantidad:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=2, column=0, sticky='w', pady=5)
    entry_cantidad = tk.Entry(frm_dialog, width=15, font=FONT_NORMAL, relief='solid', bd=1)
    entry_cantidad.grid(row=2, column=1, sticky='w', padx=5)
    entry_cantidad.insert(0, valores[2])

    tk.Label(frm_dialog, text="Precio Unitario:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=3, column=0, sticky='w', pady=5)
    entry_precio = tk.Entry(frm_dialog, width=15, font=FONT_NORMAL, relief='solid', bd=1)
    entry_precio.grid(row=3, column=1, sticky='w', padx=5)
    entry_precio.insert(0, valores[3])

    # --- Variable para almacenar resultado ---
    result = {}

    def on_ok():
        try:
            concepto = combo_concepto.get().strip()
            unidad = entry_unidad.get().strip()
            cantidad = float(entry_cantidad.get().strip())
            precio = float(entry_precio.get().strip())

            if not concepto or not unidad or cantidad <= 0 or precio < 0:
                messagebox.showwarning("Datos inválidos", "Por favor, complete todos los campos correctamente.", parent=dialog)
                return
            
            result['concepto'] = concepto
            result['unidad'] = unidad
            result['cantidad'] = cantidad
            result['precio'] = precio
            dialog.destroy()

        except ValueError:
            messagebox.showerror("Error de formato", "Cantidad y Precio deben ser números válidos.", parent=dialog)

    def on_cancel():
        dialog.destroy()

    # --- Botones ---
    frm_buttons = tk.Frame(frm_dialog, bg=BG_COLOR)
    frm_buttons.grid(row=4, column=0, columnspan=2, pady=15)

    btn_ok = tk.Button(frm_buttons, text="Aceptar", command=on_ok, bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
    btn_ok.pack(side='left', padx=10)
    btn_cancel = tk.Button(frm_buttons, text="Cancelar", command=on_cancel, bg=DANGER_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
    btn_cancel.pack(side='left', padx=10)

    root.wait_window(dialog)

    if result:
        subtotal_item = result['cantidad'] * result['precio']
        tree_items.item(item_id, values=(
            result['concepto'], 
            result['unidad'], 
            result['cantidad'], 
            f"{result['precio']:.2f}", 
            f"{subtotal_item:.2f}"
        ))
        calcular_totales_presupuesto()


def eliminar_item_presupuesto():
    sel = tree_items.selection()
    if not sel:
        messagebox.showwarning('Seleccionar', 'Selecciona un ítem para eliminar.')
        return
    
    if messagebox.askyesno("Confirmar", "¿Eliminar el ítem seleccionado?"):
        for item_id in sel:
            tree_items.delete(item_id)
        calcular_totales_presupuesto()


def limpiar_formulario_presupuesto():
    if not messagebox.askyesno("Confirmar", "¿Limpiar todo el formulario? Se perderán los datos no guardados."):
        return
    
    # Limpiar datos de cliente y obra
    entry_pres_cliente_nombre.delete(0, 'end')
    entry_pres_cliente_cuit.delete(0, 'end')
    entry_pres_cliente_domicilio.delete(0, 'end')
    entry_pres_cliente_telefono.delete(0, 'end')
    text_pres_obra_desc.delete('1.0', 'end')
    entry_pres_obra_fecha.delete(0, 'end')
    
    # Limpiar tabla de ítems
    for item in tree_items.get_children():
        tree_items.delete(item)
        
    # Limpiar condiciones y totales
    text_condiciones.delete('1.0', 'end')
    text_condiciones.insert('1.0', "Validez del presupuesto: 15 días.\nForma de pago: 50% de anticipo, 50% contra entrega.\nPlazo estimado de ejecución: a convenir.\nGarantía: 1 año sobre mano de obra.")
    
    calcular_totales_presupuesto() # Esto pondrá los totales a cero
    messagebox.showinfo("Limpiado", "Formulario listo para un nuevo presupuesto.")


def guardar_presupuesto_completo():
    # 1. Obtener y validar datos del cliente y obra
    nombre_cliente = entry_pres_cliente_nombre.get().strip()
    if not nombre_cliente:
        messagebox.showerror("Error", "El nombre del cliente es obligatorio.")
        return

    dni_cuit = entry_pres_cliente_cuit.get().strip()
    domicilio = entry_pres_cliente_domicilio.get().strip()
    telefono = entry_pres_cliente_telefono.get().strip()
    obra_desc = text_pres_obra_desc.get('1.0', 'end').strip()
    fecha_inicio = entry_pres_obra_fecha.get().strip()
    condiciones = text_condiciones.get('1.0', 'end').strip()

    if not tree_items.get_children():
        messagebox.showerror("Error", "El presupuesto debe tener al menos un ítem.")
        return

    # 2. Gestionar cliente (crear o actualizar)
    cursor.execute("SELECT id FROM clientes WHERE nombre = ?", (nombre_cliente,))
    cliente_existente = cursor.fetchone()
    
    if cliente_existente:
        cliente_id = cliente_existente[0]
        # Actualizar datos del cliente por si cambiaron
        cursor.execute("""
            UPDATE clientes 
            SET dni_cuit = ?, direccion = ?, telefono = ? 
            WHERE id = ?
        """, (dni_cuit, domicilio, telefono, cliente_id))
    else:
        # Crear nuevo cliente
        cursor.execute("""
            INSERT INTO clientes (nombre, dni_cuit, direccion, telefono) 
            VALUES (?, ?, ?, ?)
        """, (nombre_cliente, dni_cuit, domicilio, telefono))
        cliente_id = cursor.lastrowid
    
    conn.commit()
    actualizar_tree_clientes() # Refrescar la lista de clientes en su pestaña

    # 3. Guardar el presupuesto principal
    try:
        subtotal = float(entry_subtotal.get())
        iva = float(entry_iva.get())
        total = float(entry_total.get())
    except ValueError:
        messagebox.showerror("Error", "Los totales no se han calculado correctamente.")
        return

    cursor.execute("""
        INSERT INTO presupuestos (cliente_id, fecha, obra_descripcion, fecha_inicio, condiciones, subtotal, iva, monto)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (cliente_id, fecha_now(), obra_desc, fecha_inicio, condiciones, subtotal, iva, total))
    presupuesto_id = cursor.lastrowid
    
    # 4. Guardar los ítems del presupuesto
    for item_id in tree_items.get_children():
        valores = tree_items.item(item_id, 'values')
        concepto, unidad, cantidad, precio_unitario, subtotal_item = valores
        cursor.execute("""
            INSERT INTO presupuesto_items (presupuesto_id, concepto, unidad, cantidad, precio_unitario, subtotal)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (presupuesto_id, concepto, unidad, float(cantidad), float(precio_unitario), float(subtotal_item)))

    conn.commit()
    
    messagebox.showinfo("Éxito", f"Presupuesto N° {presupuesto_id} guardado correctamente.")
    limpiar_formulario_presupuesto()
    cargar_presupuestos_pendientes() # Actualizar la lista en la pestaña de cierre


def _generar_pdf_presupuesto_logic(path):
    c = canvas.Canvas(path, pagesize=A4)
    width, height = A4

    # --- Encabezado ---
    y_pos = height - 40
    
    # Dibujar logo centrado y obtener su altura
    logo_height = draw_logo_on_canvas(c, width / 2 - 50, y_pos, width=100)
    y_pos -= logo_height + 10 # Espacio después del logo

    # Datos de la empresa (centrados)
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, y_pos, obtener_config("EMPRESA_NOMBRE"))
    y_pos -= 15
    c.setFont("Helvetica", 10)
    c.drawCentredString(width / 2, y_pos, f"CUIT: {obtener_config('EMPRESA_CUIT')}")
    y_pos -= 15
    c.drawCentredString(width / 2, y_pos, obtener_config("EMPRESA_DIRECCION"))
    y_pos -= 15
    c.drawCentredString(width / 2, y_pos, f"Teléfono: {obtener_config('EMPRESA_TELEFONO')}")
    y_pos -= 15
    c.drawCentredString(width / 2, y_pos, f"Contacto: {obtener_config('EMPRESA_CONTACTO')}")
    y_pos -= 25

    # Título del documento y Fecha
    c.setFont("Helvetica-Bold", 20)
    c.drawString(40, y_pos, "Presupuesto")
    c.setFont("Helvetica", 12)
    c.drawRightString(width - 40, y_pos, f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")


    # --- Datos del Cliente y Obra ---
    y = y_pos - 30 # Ajustar 'y' para empezar más abajo
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Datos del Cliente")
    c.setFont("Helvetica", 11)
    y -= 20
    c.drawString(50, y, f"Nombre/Razón Social: {entry_pres_cliente_nombre.get().strip()}")
    y -= 15
    c.drawString(50, y, f"DNI/CUIT: {entry_pres_cliente_cuit.get().strip()}")
    y -= 15
    c.drawString(50, y, f"Domicilio: {entry_pres_cliente_domicilio.get().strip()}")
    y -= 15
    c.drawString(50, y, f"Teléfono: {entry_pres_cliente_telefono.get().strip()}")

    y -= 30
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Datos de la Obra")
    c.setFont("Helvetica", 11)
    y -= 20
    # Manejo de texto multilinea para la descripción
    desc_lines = text_pres_obra_desc.get('1.0', 'end').strip().split('\n')
    for line in desc_lines:
        c.drawString(50, y, line)
        y -= 15
    y -= 5
    c.drawString(50, y, f"Fecha de inicio prevista: {entry_pres_obra_fecha.get().strip()}")

    # --- Tabla de Ítems ---
    y -= 40
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Detalle de Materiales y Mano de Obra")
    y -= 20

    data = [['Concepto', 'Unidad', 'Cantidad', 'P. Unitario', 'Subtotal']]
    for item_id in tree_items.get_children():
        data.append(tree_items.item(item_id, 'values'))

    table = Table(data, colWidths=[260, 60, 60, 80, 80])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#DFDFDF")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'), # Alinear concepto a la izquierda
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'), # Alinear números a la derecha
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)
    
    table.wrapOn(c, width, height)
    table_height = table._height
    table.drawOn(c, 40, y - table_height)
    y -= table_height + 20

    # --- Totales ---
    c.setFont("Helvetica", 11)
    c.drawRightString(width - 40, y, f"Subtotal: ${entry_subtotal.get()}")
    y -= 20
    
    if iva_var.get():
        c.drawRightString(width - 40, y, f"IVA (21%): ${entry_iva.get()}")
        y -= 20

    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(width - 40, y, f"Total: ${entry_total.get()}")
    
    # --- Condiciones y Firma ---
    y -= 50
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Condiciones Generales")
    c.setFont("Helvetica", 10)
    y -= 20
    cond_lines = text_condiciones.get('1.0', 'end').strip().split('\n')
    for line in cond_lines:
        c.drawString(50, y, line)
        y -= 12

    # Firma
    y = 60
    c.line(width - 250, y, width - 40, y)
    c.setFont("Helvetica", 11)
    c.drawCentredString(width - 145, y - 15, "Firma del Responsable")

    c.showPage()
    c.save()

def generar_pdf_presupuesto_nuevo():
    # 1. Validar que hay datos para generar el PDF
    nombre_cliente = entry_pres_cliente_nombre.get().strip()
    if not nombre_cliente or not tree_items.get_children():
        messagebox.showerror("Error", "Debe haber un cliente y al menos un ítem para generar el PDF.")
        return

    # 2. Pedir la ruta para guardar el archivo
    path = asksaveasfilename(
        defaultextension='.pdf',
        initialdir=PDF_DIR,
        initialfile=f'Presupuesto_{nombre_cliente.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.pdf',
        filetypes=[('Archivos PDF', '*.pdf')]
    )
    if not path:
        return

    try:
        _generar_pdf_presupuesto_logic(path)
        messagebox.showinfo("Éxito", f"PDF generado correctamente en:\n{path}")
        if messagebox.askyesno("Abrir PDF", "¿Desea abrir el archivo generado?"):
            os.startfile(path)

    except Exception as e:
        messagebox.showerror("Error al generar PDF", f"Ocurrió un error: {e}")


def generar_y_enviar_email():
    # 1. Validar que hay datos
    nombre_cliente = entry_pres_cliente_nombre.get().strip()
    if not nombre_cliente or not tree_items.get_children():
        messagebox.showerror("Error", "Debe haber un cliente y al menos un ítem.")
        return

    # 2. Obtener email del cliente
    cursor.execute("SELECT email FROM clientes WHERE nombre = ?", (nombre_cliente,))
    result = cursor.fetchone()
    email_cliente = result[0] if result and result[0] else ''

    # Pedir/confirmar email
    email_destino = simpledialog.askstring("Enviar Email", "Se enviará el presupuesto a:", initialvalue=email_cliente)
    if not email_destino:
        return # User cancelled

    # 3. Generar el PDF en una ruta temporal o predefinida
    pdf_filename = f'Presupuesto_{nombre_cliente.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    path = os.path.join(PDF_DIR, pdf_filename)

    try:
        _generar_pdf_presupuesto_logic(path)
    except Exception as e:
        messagebox.showerror("Error al generar PDF", f"Ocurrió un error: {e}")
        return

    # 4. Enviar el email
    asunto = f"Presupuesto de {obtener_config('EMPRESA_NOMBRE')}"
    cuerpo = f"Estimado/a {nombre_cliente},\n\nAdjuntamos el presupuesto solicitado.\n\nSaludos cordiales,\n{obtener_config('EMPRESA_NOMBRE')}"
    
    enviar_email(email_destino, asunto, cuerpo, path)


def generar_y_enviar_whatsapp():
    # 1. Validar que hay datos
    nombre_cliente = entry_pres_cliente_nombre.get().strip()
    telefono_cliente = entry_pres_cliente_telefono.get().strip()

    if not nombre_cliente or not tree_items.get_children():
        messagebox.showerror("Error", "Debe haber un cliente y al menos un ítem.")
        return
    
    if not telefono_cliente:
        messagebox.showerror("Error", "El cliente no tiene un número de teléfono cargado.")
        return

    # 2. Generar el PDF en una ruta temporal o predefinida
    pdf_filename = f'Presupuesto_{nombre_cliente.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    path = os.path.join(PDF_DIR, pdf_filename)

    try:
        _generar_pdf_presupuesto_logic(path)
    except Exception as e:
        messagebox.showerror("Error al generar PDF", f"Ocurrió un error: {e}")
        return

    # 3. Preparar y abrir WhatsApp
    # Limpiar y formatear número de teléfono (simple, asume que ya tiene código de país)
    numero_limpio = ''.join(filter(str.isdigit, telefono_cliente))
    
    # Mensaje predefinido
    empresa_nombre = obtener_config('EMPRESA_NOMBRE')
    mensaje = f"Estimado/a {nombre_cliente},\n\nLe adjuntamos el presupuesto solicitado.\n\nSaludos cordiales,\n{empresa_nombre}"
    mensaje_url = urllib.parse.quote(mensaje)

    # Construir URL
    url = f"https://web.whatsapp.com/send?phone={numero_limpio}&text={mensaje_url}"

    # 4. Abrir navegador y notificar al usuario
    try:
        webbrowser.open(url)
        messagebox.showinfo(
            "Acción Requerida",
            f"Se ha abierto WhatsApp en su navegador.\n\n"
            f"Por favor, adjunte manualmente el siguiente archivo a la conversación:\n\n"
            f"{os.path.abspath(path)}"
        )
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el navegador: {e}")


# --- Frame superior: Datos Cliente y Obra ---
frm_pres_datos = tk.Frame(pres_main_frame, bg=BG_COLOR)
frm_pres_datos.grid(row=0, column=0, sticky='ew', padx=10, pady=5)

# Datos del cliente
frm_cliente = tk.LabelFrame(frm_pres_datos, text="Datos del Cliente", bg=BG_COLOR, fg=TEXT_COLOR, font=FONT_BOLD, padx=10, pady=10)
frm_cliente.grid(row=0, column=0, padx=5, pady=5, sticky='ns')

tk.Label(frm_cliente, text="Seleccionar Cliente:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=0, column=0, sticky='w', pady=2)
clientes_nombres = [c[1] for c in obtener_clientes()]
combo_pres_cliente_seleccionar = ttk.Combobox(frm_cliente, values=clientes_nombres, width=38, font=FONT_NORMAL)
combo_pres_cliente_seleccionar.grid(row=0, column=1, sticky='w', padx=5)
combo_pres_cliente_seleccionar.bind("<<ComboboxSelected>>", autocompletar_cliente)

tk.Label(frm_cliente, text="Nombre/Razón Social:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=1, column=0, sticky='w', pady=2)
entry_pres_cliente_nombre = tk.Entry(frm_cliente, width=40, font=FONT_NORMAL, relief='solid', bd=1)
entry_pres_cliente_nombre.grid(row=1, column=1, sticky='w', padx=5)

tk.Label(frm_cliente, text="DNI/CUIT:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=2, column=0, sticky='w', pady=2)
entry_pres_cliente_cuit = tk.Entry(frm_cliente, width=20, font=FONT_NORMAL, relief='solid', bd=1)
entry_pres_cliente_cuit.grid(row=2, column=1, sticky='w', padx=5)

tk.Label(frm_cliente, text="Domicilio:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=3, column=0, sticky='w', pady=2)
entry_pres_cliente_domicilio = tk.Entry(frm_cliente, width=50, font=FONT_NORMAL, relief='solid', bd=1)
entry_pres_cliente_domicilio.grid(row=3, column=1, sticky='w', padx=5)

tk.Label(frm_cliente, text="Teléfono/Contacto:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=4, column=0, sticky='w', pady=2)
entry_pres_cliente_telefono = tk.Entry(frm_cliente, width=30, font=FONT_NORMAL, relief='solid', bd=1)
entry_pres_cliente_telefono.grid(row=4, column=1, sticky='w', padx=5)

# Datos de la obra
frm_obra = tk.LabelFrame(frm_pres_datos, text="Datos de la Obra", bg=BG_COLOR, fg=TEXT_COLOR, font=FONT_BOLD, padx=10, pady=10)
frm_obra.grid(row=0, column=1, padx=5, pady=5, sticky='ns')

tk.Label(frm_obra, text="Descripción del Trabajo:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=0, column=0, sticky='w', pady=2)
text_pres_obra_desc = tk.Text(frm_obra, width=50, height=4, font=FONT_NORMAL, relief='solid', bd=1)
text_pres_obra_desc.grid(row=1, column=0, sticky='w', padx=5)

tk.Label(frm_obra, text="Fecha de Inicio Prevista:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=2, column=0, sticky='w', pady=2)
entry_pres_obra_fecha = tk.Entry(frm_obra, width=20, font=FONT_NORMAL, relief='solid', bd=1)
entry_pres_obra_fecha.grid(row=3, column=0, sticky='w', padx=5)

# --- Frame medio: Detalle de Materiales y Mano de Obra ---
frm_pres_items = tk.LabelFrame(pres_main_frame, text="Detalle de Materiales y Mano de Obra", bg=BG_COLOR, fg=TEXT_COLOR, font=FONT_BOLD, padx=10, pady=10)
frm_pres_items.grid(row=1, column=0, sticky='nsew', padx=10, pady=5)

cols_items = ('concepto', 'unidad', 'cantidad', 'precio_unitario', 'subtotal')
tree_items = ttk.Treeview(frm_pres_items, columns=cols_items, show='headings', height=8)
tree_items.heading('concepto', text='Concepto')
tree_items.heading('unidad', text='Unidad')
tree_items.heading('cantidad', text='Cantidad')
tree_items.heading('precio_unitario', text='Precio Unitario')
tree_items.heading('subtotal', text='Subtotal')
tree_items.column('concepto', width=300)
tree_items.column('unidad', width=80, anchor='center')
tree_items.column('cantidad', width=80, anchor='e')
tree_items.column('precio_unitario', width=120, anchor='e')
tree_items.column('subtotal', width=120, anchor='e')
tree_items.pack(side='left', fill='both', expand=True)

scroll_items = ttk.Scrollbar(frm_pres_items, orient='vertical', command=tree_items.yview)
tree_items.configure(yscrollcommand=scroll_items.set)
scroll_items.pack(side='right', fill='y')

# Botones para la tabla
frm_items_botones = tk.Frame(frm_pres_items, bg=BG_COLOR)
frm_items_botones.pack(side='left', fill='y', padx=5)

btn_item_add = tk.Button(frm_items_botones, text='Agregar', bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_item_add.pack(fill='x', pady=4)
btn_item_edit = tk.Button(frm_items_botones, text='Modificar', bg=PRIMARY_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_item_edit.pack(fill='x', pady=4)
btn_item_del = tk.Button(frm_items_botones, text='Eliminar', bg=DANGER_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_item_del.pack(fill='x', pady=4)

# --- Frame inferior: Cálculos y Condiciones ---
frm_pres_calculos = tk.Frame(pres_main_frame, bg=BG_COLOR)
frm_pres_calculos.grid(row=2, column=0, sticky='ew', padx=10, pady=5)

# Condiciones
frm_condiciones = tk.LabelFrame(frm_pres_calculos, text="Condiciones Generales", bg=BG_COLOR, fg=TEXT_COLOR, font=FONT_BOLD, padx=10, pady=10)
frm_condiciones.grid(row=0, column=0, sticky='ns')
text_condiciones = tk.Text(frm_condiciones, width=80, height=6, font=FONT_NORMAL, relief='solid', bd=1)
text_condiciones.pack(fill='both', expand=True)
text_condiciones.insert('1.0', "Validez del presupuesto: 15 días.\nForma de pago: 50% de anticipo, 50% contra entrega.\nPlazo estimado de ejecución: a convenir.\nGarantía: 1 año sobre mano de obra.")

# Totales
frm_totales = tk.LabelFrame(frm_pres_calculos, text="Totales", bg=BG_COLOR, fg=TEXT_COLOR, font=FONT_BOLD, padx=10, pady=10)
frm_totales.grid(row=0, column=1, sticky='ns', padx=10)

# Variable y Checkbox para el IVA
iva_var = tk.BooleanVar(value=True)
chk_iva = tk.Checkbutton(frm_totales, text="Aplicar IVA (21%)", variable=iva_var, bg=BG_COLOR, font=FONT_NORMAL, command=calcular_totales_presupuesto)
chk_iva.grid(row=0, column=0, columnspan=2, sticky='w', pady=3)

tk.Label(frm_totales, text="Subtotal:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=1, column=0, sticky='w', pady=3)
entry_subtotal = tk.Entry(frm_totales, width=15, font=FONT_BOLD, relief='solid', bd=1, state='readonly', justify='right')
entry_subtotal.grid(row=1, column=1, sticky='e', padx=5)

tk.Label(frm_totales, text="IVA:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=2, column=0, sticky='w', pady=3)
entry_iva = tk.Entry(frm_totales, width=15, font=FONT_BOLD, relief='solid', bd=1, state='readonly', justify='right')
entry_iva.grid(row=2, column=1, sticky='e', padx=5)

tk.Label(frm_totales, text="Total:", bg=BG_COLOR, font=FONT_BOLD).grid(row=3, column=0, sticky='w', pady=3)
entry_total = tk.Entry(frm_totales, width=15, font=FONT_BOLD, relief='solid', bd=1, state='readonly', justify='right')
entry_total.grid(row=3, column=1, sticky='e', padx=5)

# --- Frame de acciones finales ---
frm_pres_acciones = tk.Frame(pres_main_frame, bg=BG_COLOR)
frm_pres_acciones.grid(row=3, column=0, sticky='ew', padx=10, pady=10)

btn_guardar_pres = tk.Button(frm_pres_acciones, text='Guardar Presupuesto', bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=15, pady=5)
btn_guardar_pres.pack(side='right', padx=5)

btn_pdf_pres = tk.Button(frm_pres_acciones, text='Generar PDF', bg=PRIMARY_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=15, pady=5)
btn_pdf_pres.pack(side='right', padx=5)

btn_email_pres = tk.Button(frm_pres_acciones, text='Enviar por Email', bg=INFO_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=15, pady=5)
btn_email_pres.pack(side='right', padx=5)

btn_whatsapp_pres = tk.Button(frm_pres_acciones, text='Enviar por WhatsApp', bg='#25D366', fg='white', font=FONT_BOLD, relief='flat', padx=15, pady=5)
btn_whatsapp_pres.pack(side='right', padx=5)

btn_limpiar_pres = tk.Button(frm_pres_acciones, text='Limpiar Formulario', bg=WARNING_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=15, pady=5)
btn_limpiar_pres.pack(side='left', padx=5)

# Asignar comandos
btn_item_add.config(command=agregar_item_presupuesto)
btn_item_edit.config(command=modificar_item_presupuesto)
btn_item_del.config(command=eliminar_item_presupuesto)
btn_guardar_pres.config(command=guardar_presupuesto_completo)
btn_pdf_pres.config(command=generar_pdf_presupuesto_nuevo)
btn_limpiar_pres.config(command=limpiar_formulario_presupuesto)
btn_email_pres.config(command=generar_y_enviar_email)
btn_whatsapp_pres.config(command=generar_y_enviar_whatsapp)

# ---------- TAB EJECUCIÓN / CIERRE DE PRESUPUESTO ----------
frame_cierre = ttk.Frame(notebook, style="TFrame")
notebook.add(frame_cierre, text='Ejecución / Cierre')

# --- Contenido ---
cierre_top_frame = tk.Frame(frame_cierre, bg=BG_COLOR)
cierre_top_frame.pack(fill='x', pady=6, padx=10)

lbl_cierre_title = tk.Label(cierre_top_frame, text='Ejecución y Cierre de Presupuesto', font=FONT_H2, bg=BG_COLOR, fg=TEXT_COLOR)
lbl_cierre_title.pack(side='left')

# --- Formulario de Selección y Cierre ---
cierre_form_frame = tk.Frame(frame_cierre, bg=BG_COLOR, padx=10, pady=10)
cierre_form_frame.pack(fill='x')

tk.Label(cierre_form_frame, text="Presupuesto N°:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=0, column=0, sticky='w', pady=5)
combo_cierre_presupuesto = ttk.Combobox(cierre_form_frame, width=50, font=FONT_NORMAL, state="readonly")
combo_cierre_presupuesto.grid(row=0, column=1, sticky='w', padx=5)

tk.Label(cierre_form_frame, text="Cliente:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=1, column=0, sticky='w', pady=5)
entry_cierre_cliente = tk.Entry(cierre_form_frame, width=52, font=FONT_NORMAL, relief='solid', bd=1, state='readonly')
entry_cierre_cliente.grid(row=1, column=1, sticky='w', padx=5)

tk.Label(cierre_form_frame, text="Fecha de Ejecución:", bg=BG_COLOR, font=FONT_NORMAL).grid(row=2, column=0, sticky='w', pady=5)
entry_cierre_fecha = tk.Entry(cierre_form_frame, width=52, font=FONT_NORMAL, relief='solid', bd=1)
entry_cierre_fecha.grid(row=2, column=1, sticky='w', padx=5)
entry_cierre_fecha.insert(0, datetime.now().strftime('%Y-%m-%d'))

tk.Label(cierre_form_frame, text="Monto Total Cobrado:", bg=BG_COLOR, font=FONT_BOLD).grid(row=3, column=0, sticky='w', pady=5)
entry_cierre_monto = tk.Entry(cierre_form_frame, width=20, font=FONT_NORMAL, relief='solid', bd=1)
entry_cierre_monto.grid(row=3, column=1, sticky='w', padx=5)

# --- Treeview para materiales ---
cierre_tree_frame = tk.LabelFrame(frame_cierre, text="Materiales Realmente Utilizados", bg=BG_COLOR, fg=TEXT_COLOR, font=FONT_BOLD, padx=10, pady=10)
cierre_tree_frame.pack(fill='both', expand=True, padx=10, pady=10)

cols_cierre = ('material', 'cantidad_usada')
tree_cierre_materiales = ttk.Treeview(cierre_tree_frame, columns=cols_cierre, show='headings', height=6)
tree_cierre_materiales.heading('material', text='Material')
tree_cierre_materiales.heading('cantidad_usada', text='Cantidad Usada')
tree_cierre_materiales.column('material', width=400)
tree_cierre_materiales.column('cantidad_usada', width=150, anchor='e')
tree_cierre_materiales.pack(fill='both', expand=True)

# --- Lógica de la pestaña de Cierre ---

def editar_cantidad_usada(event):
    """Maneja el doble clic para editar la cantidad de un material."""
    region = tree_cierre_materiales.identify_region(event.x, event.y)
    if region != "cell":
        return

    column = tree_cierre_materiales.identify_column(event.x)
    # Permitir editar solo la columna 'cantidad_usada' (que es la #2)
    if column != '#2':
        return

    item_id = tree_cierre_materiales.identify_row(event.y)
    
    x, y, width, height = tree_cierre_materiales.bbox(item_id, column)

    # Crear un widget Entry temporal sobre la celda
    entry_edit = ttk.Entry(tree_cierre_materiales)
    
    current_value = tree_cierre_materiales.set(item_id, column)
    entry_edit.insert(0, current_value)
    entry_edit.select_range(0, 'end')
    entry_edit.focus()

    def on_save_edit(event):
        new_value = entry_edit.get()
        try:
            # Validar que el nuevo valor es un número
            float(new_value)
            tree_cierre_materiales.set(item_id, column, new_value)
        except ValueError:
            # Si no es válido, no hacer nada y simplemente destruir el Entry
            pass
        finally:
            entry_edit.destroy()

    entry_edit.bind('<Return>', on_save_edit)
    entry_edit.bind('<FocusOut>', on_save_edit)
    
    entry_edit.place(x=x, y=y, width=width, height=height)

tree_cierre_materiales.bind('<Double-1>', editar_cantidad_usada)


def cargar_presupuestos_pendientes():
    cursor.execute("""
        SELECT p.id, c.nombre, p.monto 
        FROM presupuestos p
        JOIN clientes c ON p.cliente_id = c.id
        WHERE p.estado = 'Pendiente'
        ORDER BY p.id DESC
    """)
    presupuestos = cursor.fetchall()
    combo_cierre_presupuesto['values'] = [f"N° {p[0]} - {p[1]} (Total: ${p[2]:.2f})" for p in presupuestos]

def on_presupuesto_seleccionado(event):
    seleccion = combo_cierre_presupuesto.get()
    if not seleccion:
        return

    presupuesto_id = int(seleccion.split(' ')[1])

    # Autocompletar cliente y monto
    cursor.execute("SELECT c.nombre, p.monto FROM presupuestos p JOIN clientes c ON p.cliente_id = c.id WHERE p.id = ?", (presupuesto_id,))
    data = cursor.fetchone()
    if data:
        cliente_nombre, monto_presupuestado = data
        entry_cierre_cliente.config(state='normal')
        entry_cierre_cliente.delete(0, 'end')
        entry_cierre_cliente.insert(0, cliente_nombre)
        entry_cierre_cliente.config(state='readonly')

        entry_cierre_monto.delete(0, 'end')
        entry_cierre_monto.insert(0, f"{monto_presupuestado:.2f}")

    # Cargar materiales del presupuesto en el treeview
    for item in tree_cierre_materiales.get_children():
        tree_cierre_materiales.delete(item)

    cursor.execute("""
        SELECT pi.concepto, pi.cantidad 
        FROM presupuesto_items pi
        WHERE pi.presupuesto_id = ?
    """, (presupuesto_id,))
    
    for concepto, cantidad in cursor.fetchall():
        tree_cierre_materiales.insert('', 'end', values=(concepto, cantidad))

combo_cierre_presupuesto.bind("<<ComboboxSelected>>", on_presupuesto_seleccionado)

def confirmar_cierre_presupuesto():
    seleccion = combo_cierre_presupuesto.get()
    if not seleccion:
        messagebox.showerror("Error", "Debe seleccionar un presupuesto.")
        return

    try:
        monto_cobrado = float(entry_cierre_monto.get())
    except ValueError:
        messagebox.showerror("Error", "El monto total cobrado debe ser un número válido.")
        return

    if not messagebox.askyesno("Confirmar Cierre", "¿Está seguro de que desea cerrar este presupuesto? Esta acción descontará el stock y registrará el ingreso."):
        return

    presupuesto_id = int(seleccion.split(' ')[1])

    # 1. Descontar stock
    for item_id in tree_cierre_materiales.get_children():
        material, cantidad_usada = tree_cierre_materiales.item(item_id, 'values')
        try:
            cantidad_usada = float(cantidad_usada)
            cursor.execute("UPDATE stock SET cantidad = cantidad - ? WHERE producto = ?", (cantidad_usada, material))
        except ValueError:
            messagebox.showwarning("Advertencia", f"La cantidad para '{material}' no es un número válido. No se descontará del stock.")
            continue
    
    # 2. Registrar ingreso
    descripcion_ingreso = f"Pago cliente - Presupuesto N° {presupuesto_id}"
    fecha_ingreso = entry_cierre_fecha.get()
    cursor.execute("INSERT INTO movimientos (tipo, descripcion, monto, fecha) VALUES (?, ?, ?, ?)",
                   ('INGRESO', descripcion_ingreso, monto_cobrado, fecha_ingreso))

    # 3. Actualizar estado del presupuesto
    cursor.execute("UPDATE presupuestos SET estado = 'Concretado' WHERE id = ?", (presupuesto_id,))

    conn.commit()

    # 4. Actualizar Vistas
    actualizar_tree_stock()
    actualizar_tree_movimientos()
    cargar_presupuestos_pendientes() # Recargar la lista de pendientes

    # 5. Limpiar formulario y notificar
    entry_cierre_cliente.config(state='normal')
    entry_cierre_cliente.delete(0, 'end')
    entry_cierre_cliente.config(state='readonly')
    entry_cierre_monto.delete(0, 'end')
    entry_cierre_fecha.delete(0, 'end')
    entry_cierre_fecha.insert(0, datetime.now().strftime('%Y-%m-%d'))
    combo_cierre_presupuesto.set('')
    for item in tree_cierre_materiales.get_children():
        tree_cierre_materiales.delete(item)

    messagebox.showinfo("Éxito", f"Presupuesto N° {presupuesto_id} cerrado y registrado correctamente.")


# --- Botón de Confirmación ---
btn_confirmar_cierre = tk.Button(frame_cierre, text='Confirmar Cierre', command=confirmar_cierre_presupuesto, bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=15, pady=5)
btn_confirmar_cierre.pack(pady=10)

# Cargar presupuestos pendientes al iniciar
cargar_presupuestos_pendientes()


# ---------- TAB MOVIMIENTOS (INGRESOS / EGRESOS / SALDO) ----------
frame_mov = ttk.Frame(notebook, style="TFrame")
notebook.add(frame_mov, text='Ingresos / Egresos')

frm_mov_top = tk.Frame(frame_mov, bg=BG_COLOR)
frm_mov_top.pack(fill='x', pady=6)

lbl_mov = tk.Label(frm_mov_top, text='Registrar Movimiento (Manualmente)', font=FONT_H2, bg=BG_COLOR, fg=TEXT_COLOR)
lbl_mov.pack(side='left', padx=6)

frm_mov_form = tk.Frame(frame_mov, bg=BG_COLOR)
frm_mov_form.pack(fill='x', padx=8, pady=6)

tk.Label(frm_mov_form, text='Tipo:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=0, column=0, sticky='w')
combo_mov_tipo = ttk.Combobox(frm_mov_form, values=['INGRESO','EGRESO'], width=15)
combo_mov_tipo.grid(row=0, column=1, padx=6, pady=2)
combo_mov_tipo.current(0)

tk.Label(frm_mov_form, text='Descripcion:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=1, column=0, sticky='w')
entry_mov_desc = tk.Entry(frm_mov_form, width=60, font=FONT_NORMAL, relief='solid', bd=1)
entry_mov_desc.grid(row=1, column=1, padx=6, pady=2)

tk.Label(frm_mov_form, text='Monto:', bg=BG_COLOR, font=FONT_NORMAL).grid(row=2, column=0, sticky='w')
entry_mov_monto = tk.Entry(frm_mov_form, width=20, font=FONT_NORMAL, relief='solid', bd=1)
entry_mov_monto.grid(row=2, column=1, sticky='w', pady=2)


def agregar_movimiento():
    tipo = combo_mov_tipo.get()
    desc = entry_mov_desc.get().strip()
    try:
        monto = float(entry_mov_monto.get().strip())
    except Exception:
        messagebox.showwarning('Error', 'Monto inválido')
        return
    fecha = fecha_now()
    cursor.execute('INSERT INTO movimientos(tipo, descripcion, monto, fecha) VALUES(?,?,?,?)', (tipo, desc, monto, fecha))
    conn.commit()
    entry_mov_desc.delete(0, 'end')
    entry_mov_monto.delete(0, 'end')
    actualizar_tree_movimientos()

btn_mov_add = tk.Button(frm_mov_form, text='Registrar Movimiento', command=agregar_movimiento, bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=4)
btn_mov_add.grid(row=0, column=2, rowspan=3, padx=8)

# Tree movimientos
cols_mov = ('id', 'tipo', 'descripcion', 'monto', 'fecha')
tree_mov = ttk.Treeview(frame_mov, columns=cols_mov, show='headings', height=12)
for c in cols_mov:
    tree_mov.heading(c, text=c.capitalize())
tree_mov.pack(fill='both', padx=8, pady=6)


def actualizar_tree_movimientos():
    for r in tree_mov.get_children():
        tree_mov.delete(r)
    cursor.execute('SELECT id, tipo, descripcion, monto, fecha FROM movimientos ORDER BY fecha DESC')
    total = 0.0
    for row in cursor.fetchall():
        tree_mov.insert('', 'end', values=row)
        if row[1] == 'INGRESO':
            total += float(row[3])
        else:
            total -= float(row[3])
    lbl_saldo.config(text=f"Saldo actual: ${total:.2f}")

# saldo label and export
lbl_saldo = tk.Label(frame_mov, text='Saldo actual: $0.00', font=(FONT_FAMILY, 12, "bold"), bg=BG_COLOR, fg=TEXT_COLOR)
lbl_saldo.pack(pady=6)

btn_export_balance = tk.Button(frame_mov, text='Exportar Balance a PDF', command=lambda: (generar_pdf_reporte_balance(asksaveasfilename(defaultextension='.pdf', initialfile='Balance.pdf', filetypes=[('PDF','*.pdf')] ) or None), messagebox.showinfo('Exportado','Balance exportado') ), bg=PRIMARY_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=5)
btn_export_balance.pack(pady=6)

actualizar_tree_movimientos()

# ---------- TAB CONFIGURACION ----------
frame_config = ttk.Frame(notebook, style="TFrame")
notebook.add(frame_config, text='Configuración')

frm_conf_top = tk.Frame(frame_config, bg=BG_COLOR)
frm_conf_top.pack(fill='x', pady=6)

lbl_conf = tk.Label(frm_conf_top, text='Datos de la Empresa para PDF', font=FONT_H2, bg=BG_COLOR, fg=TEXT_COLOR)
lbl_conf.pack(side='left', padx=6)

frm_conf_form = tk.Frame(frame_config, bg=BG_COLOR, padx=8, pady=6)
frm_conf_form.pack(fill='x')

# Campos del formulario
labels_config = {
    "EMPRESA_NOMBRE": "Nombre de la Empresa:",
    "EMPRESA_CUIT": "CUIT:",
    "EMPRESA_DIRECCION": "Dirección:",
    "EMPRESA_TELEFONO": "Teléfono:",
    "EMPRESA_CONTACTO": "Email/Contacto:"
}

entries_config = {}
for i, (clave, texto) in enumerate(labels_config.items()):
    tk.Label(frm_conf_form, text=texto, bg=BG_COLOR, font=FONT_NORMAL).grid(row=i, column=0, sticky='w', padx=5, pady=5)
    entry = tk.Entry(frm_conf_form, width=60, font=FONT_NORMAL, relief='solid', bd=1)
    entry.grid(row=i, column=1, sticky='w', padx=5, pady=5)
    entries_config[clave] = entry

# --- SMTP Config ---
frm_smtp_form = tk.LabelFrame(frame_config, text="Configuración de Email (SMTP)", bg=BG_COLOR, fg=TEXT_COLOR, font=FONT_BOLD, padx=8, pady=6)
frm_smtp_form.pack(fill='x', padx=8, pady=10)

labels_smtp = {
    "SMTP_SERVER": "Servidor SMTP:",
    "SMTP_PORT": "Puerto:",
    "SMTP_USER": "Usuario (Email):",
    "SMTP_PASS": "Contraseña:"
}

entries_smtp = {}
for i, (clave, texto) in enumerate(labels_smtp.items()):
    tk.Label(frm_smtp_form, text=texto, bg=BG_COLOR, font=FONT_NORMAL).grid(row=i, column=0, sticky='w', padx=5, pady=5)
    entry = tk.Entry(frm_smtp_form, width=60, font=FONT_NORMAL, relief='solid', bd=1)
    if 'PASS' in clave:
        entry.config(show='*')
    entry.grid(row=i, column=1, sticky='w', padx=5, pady=5)
    entries_smtp[clave] = entry

def cargar_datos_configuracion():
    all_entries = {**entries_config, **entries_smtp}
    for clave, entry in all_entries.items():
        entry.delete(0, 'end')
        entry.insert(0, obtener_config(clave))

def guardar_datos_configuracion():
    all_entries = {**entries_config, **entries_smtp}
    for clave, entry in all_entries.items():
        guardar_config(clave, entry.get().strip())
    messagebox.showinfo("Guardado", "Los datos de la empresa se han guardado correctamente.")

btn_conf_guardar = tk.Button(frame_config, text='Guardar Cambios', command=guardar_datos_configuracion, bg=SUCCESS_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=5)
btn_conf_guardar.pack(pady=20)

# Cargar datos al iniciar la pestaña
cargar_datos_configuracion()

def seleccionar_logo():
    global LOGO_PATH
    filepath = filedialog.askopenfilename(
        title="Seleccionar archivo de logo",
        filetypes=[("Imágenes", "*.png *.jpg *.jpeg"), ("Todos los archivos", "*.*")]
    )
    if not filepath:
        return

    try:
        # Usar un nombre de archivo fijo para el logo en la carpeta de la app
        destination_path = "logo_empresa.png"
        shutil.copy(filepath, destination_path)
        guardar_config("EMPRESA_LOGO", destination_path)
        LOGO_PATH = destination_path # Actualizar la variable global
        messagebox.showinfo("Éxito", "Logo actualizado correctamente. Se reflejará en los próximos PDFs.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el logo: {e}")

btn_conf_logo = tk.Button(frame_config, text='Cargar Logo de Empresa', command=seleccionar_logo, bg=INFO_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=5)
btn_conf_logo.pack(pady=10)


# ---------- BOTONES GLOBALES ----------
frm_global = tk.Frame(root, bg=BG_COLOR)
frm_global.pack(fill='x', padx=10, pady=6)

def actualizar_tree_presupuestos():
    # Esta función está pendiente de implementación.
    # Se deja como placeholder para evitar errores.
    pass

btn_refresh = tk.Button(frm_global, text='Refrescar todo', command=lambda: (actualizar_tree_clientes(), actualizar_tree_proveedores(), actualizar_tree_stock(), actualizar_tree_presupuestos(), actualizar_tree_movimientos(), cargar_presupuestos_pendientes()), bg=INFO_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=5)
btn_refresh.pack(side='left', padx=6)

btn_salir = tk.Button(frm_global, text='Salir', command=root.destroy, bg=DANGER_COLOR, fg='white', font=FONT_BOLD, relief='flat', padx=10, pady=5)
btn_salir.pack(side='right', padx=6)

# Ejecutar mainloop
root.mainloop()
